from bs4 import BeautifulSoup
import os
import pandas as pd
from openpyxl import load_workbook

dir_path = os.path.dirname(os.path.realpath(__file__))
meas = pd.DataFrame(columns=['accuracy', 'cross', 'n_vars', 'variables'])
index = 0

#For Analysis Results
folders = [r"/Candidates2Params(NonFiltered)", r"/Candidates2Params(NonFiltered,pucalculated)",
          r"/Candidates2Params(v1-v2, NonFiltered, pucalculated)", r"/Candidates2Params(50HzFiltered)",
          r"/Candidates2Params(50HZFiltered,pucalculated)", r"/Candidates2Params(v1-v2,50HzFiltered, pucalculated)"]
studies = [r"/Wilks_ICS.htm", r"/Wilks_delta.htm", r"/Mahalanobis_ICS.htm", r"/Mahalanobis_ICS.htm"]

##For Josechu Comparisons
# folders = [r"/Results_SPSS/6VarsJosechuDelta/6Vars_Josechu_delta"]
# studies = [r"(NonFiltered).htm", r"(NonFiltered, pucalculated).htm", r"(v1-v2, NonFiltered, pucalculated).htm",
#            r"(50HzFiltered).htm", r"(50HzFiltered, pucalculated).htm", r"(v1-v2, 50HzFiltered, pucalculated).htm"]



for fol in folders:
    print("--------"+fol[19:-1]+"--------")
    for stu in studies:
        print(stu[1:-4]+": ", end='')
        with open(dir_path + fol + stu) as f:
            html = f.read()
            soup = BeautifulSoup(html, "lxml")
            x = soup.find_all('table')
            y = x[len(x) - 1]
            z = y.find_all('td')
            for lines in z:
                if str(lines).find('cross-validated grouped cases correctly classified') != -1:
                    cross = str(lines)[37:42]
                elif str(lines).find('original grouped cases correctly classified') != -1:
                    res = str(lines)[37:42]
            i = 0
            n = -1
            t = 0
            variables = []
            for j in soup.find_all('caption'):
                n = n + 1
                if "Standardized Canonical Discriminant Function Coefficients" in j:
                    for row in soup.find_all('table')[n].tbody.findAll('tr'):
                        i = i + 1
                        if i >= 3 and str(row.findAll('td')[0].contents)!='[]':
                            t = t + 1
                            variables.append(str(row.findAll('td')[0].contents)[2:-2])
            print(
                '{} and {} of cross correlation with {} variables used being: {}'.format(res, cross, t, variables))
            meas.loc[index] = [res, cross, t,
                               str(variables)[1:-1].replace(", ", "\t").replace(r"'", r"\"").replace('\\', "")]
            index = index + 1

#book = load_workbook(r'C:\Users\Jorge\GDrive\Universidad\BSICOS\Project\results_plantilla_josechu.xlsx')
book = load_workbook(r'C:\Users\Jorge\GDrive\Universidad\BSICOS\Project\results_plantilla.xlsx')
writer = pd.ExcelWriter(r'C:\Users\Jorge\GDrive\Universidad\BSICOS\Project\Results_SPSS\ResultingVars_ExtraBigRange\results.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
meas.to_excel(writer, sheet_name='Sheet1')
writer.save()
